import os
import json
import ast
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
import sys
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
if project_root not in sys.path:
    sys.path.append(project_root)

def is_long_text(value):
    return isinstance(value, str) and len(value) > 150


def get_code_and_evidence(cb):
    if 'code' in cb:
        return cb['code'], cb['evidence']
    elif '*code' in cb:
        return f"*{cb['*code']}", f"*{cb['*evidence']}"
    else:
        return '', ''


def merge_same_cells(ws, col_name):
    col_idx = None
    for i, cell in enumerate(ws[1], 1):
        if cell.value == col_name:
            col_idx = i
            break
    if col_idx is None:
        return
    start_row = 2
    current_value = ws.cell(row=start_row, column=col_idx).value
    merge_start = start_row
    for row in range(start_row + 1, ws.max_row + 1):
        value = ws.cell(row=row, column=col_idx).value
        if value != current_value:
            if row - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                ws.cell(row=merge_start, column=col_idx).alignment = Alignment(vertical='center', horizontal='center',
                                                                               wrap_text=True)
            current_value = value
            merge_start = row
    if ws.max_row > merge_start:
        ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        ws.cell(row=merge_start, column=col_idx).alignment = Alignment(vertical='center', horizontal='center',
                                                                       wrap_text=True)


def process_json_file(file_path, output_dir):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Codebook sheet
    codebook_records = []
    for cb in data['Codebook']:
        code, evidence = get_code_and_evidence(cb)
        codebook_records.append({
            'target_text': data['target_text'],
            'code': code,
            'evidence': evidence
        })
    codebook_df = pd.DataFrame(codebook_records)

    # Consolidating_results sheet
    consolidating_records = []
    for a in data['Consolidating_results']['Agreed']:
        consolidating_records.append({
            'target_text': data['target_text'],
            'agreed/disagreed': 'Agreed',
            'code': a['code'],
            'evidence': a['evidence']
        })
    for d in data['Consolidating_results']['Disagreed']:
        consolidating_records.append({
            'target_text': data['target_text'],
            'agreed/disagreed': 'Disagreed',
            'code': d['code'],
            'evidence': d['evidence']
        })
    consolidating_df = pd.DataFrame(consolidating_records)

    # Debate sheet
    debate_records = []
    for debate in data['Debate']:
        target_text = data['target_text']
        disagreed_code = debate['Disagreed']
        for i, process in enumerate(debate['Process']):
            round_number = i + 1
            responses = ast.literal_eval(process['response']) if isinstance(process['response'], str) else process[
                'response']
            for res in responses:
                debate_records.append({
                    'Target text': target_text,
                    'Disagreed code': disagreed_code,
                    'Round': f"Round {round_number}",
                    'Response': res
                })
    debate_df = pd.DataFrame(debate_records)

    # Role_Team sheet — merge columns and reorder
    role_team_df = pd.DataFrame(data['Role_Team'])
    role_team_df['role / disciplinary_background / core_value'] = role_team_df.apply(
        lambda row: f"{row['role']} / {row['disciplinary_background']} / {row['core_value']}", axis=1
    )
    role_team_final_df = role_team_df[['role / disciplinary_background / core_value', 'positionality']]
    # Swap columns
    role_team_final_df = role_team_final_df[['role / disciplinary_background / core_value', 'positionality']]

    # Save Excel
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}_processed.xlsx")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        codebook_df.to_excel(writer, sheet_name='Codebook', index=False)
        role_team_final_df.to_excel(writer, sheet_name='Role_Team', index=False)
        consolidating_df.to_excel(writer, sheet_name='Consolidating_results', index=False)
        debate_df.to_excel(writer, sheet_name='Debate', index=False)

    wb = load_workbook(output_path)

    # Merge cells
    merge_same_cells(wb['Codebook'], 'target_text')
    merge_same_cells(wb['Consolidating_results'], 'target_text')
    merge_same_cells(wb['Consolidating_results'], 'agreed/disagreed')
    merge_same_cells(wb['Debate'], 'Target text')
    merge_same_cells(wb['Debate'], 'Disagreed code')
    merge_same_cells(wb['Debate'], 'Round')

    # Formatting
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if not is_long_text(cell.value):
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            adjusted_width = (max_length + 2) if max_length < 100 else 100
            ws.column_dimensions[column].width = adjusted_width

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            max_lines = 1
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    lines = cell.value.count("\n") + cell.value.count(" ") // 20 + 2
                    if lines > max_lines:
                        max_lines = lines
            ws.row_dimensions[cell.row].height = max_lines * 15

    wb.save(output_path)
    print(f"✅ Processed and saved: {output_path}")


# ========= 🚀 执行入口 =========
input_folder = r'F:\Work\Debate\MultiAgentDabateDataAnnotation\Data\Scrum-interviews\output\debate_process\json'  # 👈 修改为你的 JSON 文件夹路径
output_folder = r'F:\Work\Debate\MultiAgentDabateDataAnnotation\Data\Scrum-interviews\output\debate_process\excel'  # 👈 输出 Excel 文件夹路径
os.makedirs(output_folder, exist_ok=True)

# 迭代所有 JSON 文件
for filename in os.listdir(input_folder):
    if filename.endswith('.json'):
        file_path = os.path.join(input_folder, filename)
        process_json_file(file_path, output_folder)