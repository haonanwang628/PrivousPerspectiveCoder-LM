import os
import json
import ast
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sys

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
if project_root not in sys.path:
    sys.path.append(project_root)


def is_long_text(value):
    return isinstance(value, str) and len(value) > 150


def get_code_and_definition(cb):
    if 'code' in cb:
        return cb['code'], cb['definition']
    elif '*code' in cb:
        return f"*{cb['*code']}", f"*{cb['*definition']}"
    else:
        return '', ''


def process_json_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Codebook
    codebook_records = []
    for cb in data['Codebook']:
        code, definition = get_code_and_definition(cb)
        codebook_records.append({
            'target_text': data['target_text'],
            'code': code,
            'definition': definition
        })
    codebook_df = pd.DataFrame(codebook_records)

    # Consolidating_results
    consolidating_records = []
    for a in data['Consolidating_results']['Agreed']:
        consolidating_records.append({
            'target_text': data['target_text'],
            'agreed/disagreed': 'Agreed',
            'code': a['code'],
            'definition': a['definition']
        })
    for d in data['Consolidating_results']['Disagreed']:
        consolidating_records.append({
            'target_text': data['target_text'],
            'agreed/disagreed': 'Disagreed',
            'code': d['code'],
            'definition': d['definition']
        })
    consolidating_df = pd.DataFrame(consolidating_records)

    # Debate
    debate_records = []
    for debate in data['Debate']:
        target_text = data['target_text']
        disagreed_code = debate['Disagreed']

        closing_dict = debate.get('Closing', {})
        closing_str = "\n".join([f"{k}: {v}" for k, v in closing_dict.items()])

        round_responses = []
        for i, process in enumerate(debate['Process']):
            responses = ast.literal_eval(process['response']) if isinstance(process['response'], str) else process[
                'response']
            round_responses.append(responses)

        num_roles = len(round_responses[0])
        for role_idx in range(num_roles):
            row_record = {
                'Target text': target_text,
                'Disagreed code': disagreed_code
            }
            for rnd_num, responses_in_round in enumerate(round_responses):
                column_name = f'Round {rnd_num + 1}'
                row_record[column_name] = responses_in_round[role_idx]
            row_record['Closing'] = closing_str
            debate_records.append(row_record)

    debate_df = pd.DataFrame(debate_records)

    # Role_Team
    role_team_df = pd.DataFrame(data['Role_Team'])
    role_team_df['perspective'] = role_team_df.apply(
        lambda
            row: f"{row['role']} / {row['Intended_Study_Level']} / {row['Subject']} "
                 f"/ {row['Research_Interest']} /{row['Dimensions_Source']}",
        axis=1
    )
    role_team_final_df = role_team_df[['perspective', 'positionality']]
    role_team_final_df = role_team_final_df[['perspective', 'positionality']]

    return codebook_df, consolidating_df, debate_df, role_team_final_df


def merge_same_cells(ws, col_name):
    col_idx = None
    for i, cell in enumerate(ws[1], 1):
        if cell.value == col_name:
            col_idx = i
            break
    if col_idx is None:
        return
    start_row = 2
    current_value = ws.cell(row=start_row, column=col_idx).value
    merge_start = start_row
    for row in range(start_row + 1, ws.max_row + 1):
        value = ws.cell(row=row, column=col_idx).value
        if value != current_value:
            if row - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                ws.cell(row=merge_start, column=col_idx).alignment = Alignment(vertical='center', horizontal='center',
                                                                               wrap_text=True)
            current_value = value
            merge_start = row
    if ws.max_row > merge_start:
        ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        ws.cell(row=merge_start, column=col_idx).alignment = Alignment(vertical='center', horizontal='center',
                                                                       wrap_text=True)


if __name__ == "__main__":
    input_folder = r'F:\Work\Debate\MultiAgentDabateDataAnnotation\Data\Scrum-interviews\output\debate_process\json'
    output_folder = r'F:\Work\Debate\MultiAgentDabateDataAnnotation\Data\Scrum-interviews\output\debate_process\excel'
    os.makedirs(output_folder, exist_ok=True)

    total_codebook_df = pd.DataFrame()
    total_consolidating_df = pd.DataFrame()
    total_debate_df = pd.DataFrame()
    final_role_team_df = None

    for filename in os.listdir(input_folder):
        if filename.endswith('.json'):
            file_path = os.path.join(input_folder, filename)
            codebook_df, consolidating_df, debate_df, role_team_df = process_json_file(file_path)

            total_codebook_df = pd.concat([total_codebook_df, codebook_df], ignore_index=True)
            total_consolidating_df = pd.concat([total_consolidating_df, consolidating_df], ignore_index=True)
            total_debate_df = pd.concat([total_debate_df, debate_df], ignore_index=True)
            final_role_team_df = role_team_df

    output_path = os.path.join(output_folder, "merged_all_processed.xlsx")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        total_codebook_df.to_excel(writer, sheet_name='Codebook', index=False)
        final_role_team_df.to_excel(writer, sheet_name='Role_Team', index=False)
        total_consolidating_df.to_excel(writer, sheet_name='Consolidating_results', index=False)
        total_debate_df.to_excel(writer, sheet_name='Debate', index=False)

    wb = load_workbook(output_path)

    # 合并
    merge_same_cells(wb['Codebook'], 'target_text')
    merge_same_cells(wb['Consolidating_results'], 'target_text')
    merge_same_cells(wb['Consolidating_results'], 'agreed/disagreed')
    merge_same_cells(wb['Debate'], 'Target text')
    merge_same_cells(wb['Debate'], 'Disagreed code')

    # 格式化
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if not is_long_text(cell.value):
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            adjusted_width = (max_length + 2) if max_length < 100 else 100
            ws.column_dimensions[column].width = adjusted_width

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            max_lines = 1
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    lines = cell.value.count("\n") + cell.value.count(" ") // 20 + 2
                    if lines > max_lines:
                        max_lines = lines
            ws.row_dimensions[cell.row].height = max_lines * 15

    wb.save(output_path)
    print(f"✅ All JSON files merged and saved: {output_path}")
