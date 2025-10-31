import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sys

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
if project_root not in sys.path:
    sys.path.append(project_root)


# ----------------- 你的解析逻辑（保持） -----------------
def get_code_and_definition(cb):
    if 'code' in cb:
        return cb['code'], cb.get('justification', '')
    else:
        return '', ''


def process_json_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Codebook：注意此处依据你给的结构 data['Codebook'][0]['codebook']
    codebook_records = []
    for cb in data['Codebook'][0]['codebook']:
        code, definition = get_code_and_definition(cb)
        codebook_records.append({
            'target_text': data['target_text'],
            'code': code,
            'definition': definition
        })
    codebook_df = pd.DataFrame(codebook_records)

    # Role_Team：你给的新结构是 data['Role']
    role_team_df = pd.DataFrame(data['Role'])
    role_team_df['perspective'] = role_team_df.apply(
        lambda row: f"{row['role']} / {row['Intended_Study_Level']} / {row['Subject']} "
                    f"/ {row['Research_Interest']} /{row['Dimensions_Source']}",
        axis=1
    )
    role_team_final_df = role_team_df[['perspective', 'positionality']].copy()
    return codebook_df, role_team_final_df


# ----------------- 新增：格式化与合并单元格 -----------------
def _merge_same_cells(ws, col_name):
    """合并同一列相邻相同值的单元格（用于 target_text）"""
    # 找列号
    col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == col_name:
            col_idx = idx
            break
    if col_idx is None or ws.max_row < 2:
        return

    start_row = 2
    curr = ws.cell(row=start_row, column=col_idx).value
    block_start = start_row

    for r in range(start_row + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        if val != curr:
            if r - 1 >= block_start:
                ws.merge_cells(start_row=block_start, start_column=col_idx,
                               end_row=r - 1, end_column=col_idx)
                ws.cell(row=block_start, column=col_idx).alignment = Alignment(
                    vertical='center', horizontal='center', wrap_text=True
                )
            curr = val
            block_start = r

    # 末尾收尾
    if ws.max_row >= block_start:
        ws.merge_cells(start_row=block_start, start_column=col_idx,
                       end_row=ws.max_row, end_column=col_idx)
        ws.cell(row=block_start, column=col_idx).alignment = Alignment(
            vertical='center', horizontal='center', wrap_text=True
        )


def _autotune_sheet(ws):
    """自动换行/列宽/行高 & 冻结首行"""
    # 对齐与换行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            s = str(cell.value) if cell.value is not None else ""
            if len(s) > 150:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # 列宽
    for col in ws.columns:
        col_letter = col[0].column_letter
        maxlen = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(100, max(12, maxlen + 2))

    # 行高（粗略估计）
    for r in range(2, ws.max_row + 1):
        max_lines = 1
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                approx_lines = v.count("\n") + max(1, len(v) // 40)
                if approx_lines > max_lines:
                    max_lines = approx_lines
        ws.row_dimensions[r].height = max(15, max_lines * 15)

    ws.freeze_panes = "A2"


# ----------------- 主流程（在你基础上最小改动） -----------------
if __name__ == "__main__":
    input_folder = r'F:\Work\Debate\MultiAgentDabateDataAnnotation\Data\Scrum-interviews\output\baseline2\json'
    output_folder = r'F:\Work\Debate\MultiAgentDabateDataAnnotation\Data\Scrum-interviews\output\baseline2\excel'
    os.makedirs(output_folder, exist_ok=True)

    # Codebook 分角色累加；Role_Team 各自覆盖后再合并成一张表
    codebook_by_role = {1: [], 2: [], 3: []}
    role_last_df = {1: None, 2: None, 3: None}


    def _which_role(name: str):
        n = name.lower()
        if 'role1' in n: return 1
        if 'role2' in n: return 2
        if 'role3' in n: return 3
        return None


    # 稳定排序，确保“最后一个覆盖”是确定的
    for filename in sorted(os.listdir(input_folder)):
        if filename.lower().endswith('.json'):
            file_path = os.path.join(input_folder, filename)
            try:
                codebook_df, role_team_df = process_json_file(file_path)
                rid = _which_role(filename)
                if rid is not None:
                    codebook_by_role[rid].append(codebook_df.copy())  # Codebook：按角色累加（多行结构）
                    role_last_df[rid] = role_team_df.copy()  # Role_Team：按角色覆盖（仅保留最后一个）
            except Exception as e:
                print(f"[WARN] 跳过文件（解析失败）: {filename} -> {e}")

    def _stack_or_empty(lst):
        return pd.concat(lst, ignore_index=True) if lst else pd.DataFrame(columns=["target_text", "code", "definition"])


    codebook_role1 = _stack_or_empty(codebook_by_role[1])
    codebook_role2 = _stack_or_empty(codebook_by_role[2])
    codebook_role3 = _stack_or_empty(codebook_by_role[3])

    # Role_Team：合成一张表（加 role 列；把 3 个角色的最后版本竖向合并）
    role_team_rows = []
    for rid in (1, 2, 3):
        df = role_last_df[rid]
        if df is None:
            df = pd.DataFrame(columns=["perspective", "positionality"])
        df = df.copy()
        df.insert(0, "role", f"role{rid}")
        role_team_rows.append(df)
    role_team_merged = pd.concat(role_team_rows, ignore_index=True)

    # 导出一个 Excel（4 个 sheet）
    output_path = os.path.join(output_folder, "merged_all.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        role_team_merged.to_excel(writer, sheet_name="Role_Team_All", index=False)
        codebook_role1.to_excel(writer, sheet_name="Codebook_role1", index=False)
        codebook_role2.to_excel(writer, sheet_name="Codebook_role2", index=False)
        codebook_role3.to_excel(writer, sheet_name="Codebook_role3", index=False)

    # 打开进行“target_text”纵向合并 + 美化（仅对 Codebook 三表）
    wb = load_workbook(output_path)
    for sheet in ("Codebook_role1", "Codebook_role2", "Codebook_role3"):
        ws = wb[sheet]
        if ws.max_row >= 2:
            _merge_same_cells(ws, "target_text")  # ✅ 同一 target_text 的多条 code 合并成一个大单元格
            _autotune_sheet(ws)
    # Role_Team_All 也做基本美化
    _autotune_sheet(wb["Role_Team_All"])

    wb.save(output_path)
    print(f"✅ 导出完成：{output_path}")
