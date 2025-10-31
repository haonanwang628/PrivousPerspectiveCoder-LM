import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sys

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
if project_root not in sys.path:
    sys.path.append(project_root)

# --- 兼容大小写键名 ---
def get_code_and_definition(cb):
    # 支持 "Code/Definition" 或 "code/definition"
    code = cb.get("Code", cb.get("code", ""))
    definition = cb.get("Definition", cb.get("definition", ""))
    return code, definition

def process_json_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Codebook 展开成多行：target_text + code + definition
    rows = []
    for cb in data.get('Codebook', []):
        code, definition = get_code_and_definition(cb)
        rows.append({
            'target_text': data.get('target_text', ''),
            'code': code,
            'definition': definition
        })
    return pd.DataFrame(rows, columns=["target_text", "code", "definition"])

# --- 样式与合并 ---
def _merge_same_cells(ws, col_name):
    # 找到列号
    col_idx = None
    for i, cell in enumerate(ws[1], 1):
        if cell.value == col_name:
            col_idx = i
            break
    if col_idx is None or ws.max_row < 2:
        return

    start = 2
    curr = ws.cell(row=start, column=col_idx).value
    block_start = start

    for r in range(start + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        if val != curr:
            if r - 1 >= block_start:
                ws.merge_cells(start_row=block_start, start_column=col_idx, end_row=r - 1, end_column=col_idx)
                ws.cell(row=block_start, column=col_idx).alignment = Alignment(
                    vertical='center', horizontal='center', wrap_text=True
                )
            curr = val
            block_start = r

    if ws.max_row >= block_start:
        ws.merge_cells(start_row=block_start, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        ws.cell(row=block_start, column=col_idx).alignment = Alignment(
            vertical='center', horizontal='center', wrap_text=True
        )

def _autotune_sheet(ws):
    # 自动换行与对齐
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            text = str(cell.value) if cell.value is not None else ""
            if len(text) > 150:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # 列宽
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(100, max(12, max_len + 2))

    # 行高（粗估：每40字符≈1行）
    for r in range(2, ws.max_row + 1):
        max_lines = 1
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                approx = v.count("\n") + max(1, len(v) // 40)
                max_lines = max(max_lines, approx)
        ws.row_dimensions[r].height = max(15, max_lines * 15)

    ws.freeze_panes = "A2"

if __name__ == "__main__":
    input_folder = r'F:\Work\Debate\MultiAgentDabateDataAnnotation\Data\Scrum-interviews\output\baseline1\json'
    output_folder = r'F:\Work\Debate\MultiAgentDabateDataAnnotation\Data\Scrum-interviews\output\baseline1\excel'
    os.makedirs(output_folder, exist_ok=True)

    # 读入并合并
    frames = []
    for filename in sorted(os.listdir(input_folder)):  # 排序保证稳定
        if filename.lower().endswith('.json'):
            file_path = os.path.join(input_folder, filename)
            try:
                frames.append(process_json_file(file_path))
            except Exception as e:
                print(f"[WARN] 跳过文件（解析失败）: {filename} -> {e}")

    if frames:
        total_codebook_df = pd.concat(frames, ignore_index=True)
    else:
        total_codebook_df = pd.DataFrame(columns=["target_text", "code", "definition"])

    # 写出
    output_path = os.path.join(output_folder, "merged_all.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        total_codebook_df.to_excel(writer, sheet_name="Codebook", index=False)

    # 打开后做合并与美化（与“之前的excel格式”一致）
    wb = load_workbook(output_path)
    ws = wb["Codebook"]
    if ws.max_row >= 2:
        _merge_same_cells(ws, "target_text")  # 同一 target_text 的相邻行纵向合并
    _autotune_sheet(ws)
    wb.save(output_path)

    print(f"✅ 导出完成：{output_path}")
